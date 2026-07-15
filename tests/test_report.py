"""
Unit tests for the Portfolio Report Generator module.
"""

import os
import tempfile
from datetime import datetime, timedelta
import pytest
import openpyxl

from portfolio.importer import Trade
from portfolio.report import PortfolioReportGenerator
from portfolio.risk import RiskReport, RiskParameters


@pytest.fixture
def mock_backtest_data():
    """Generates synthetic backtest results, trades, and conflict logs."""
    t = datetime(2026, 1, 1, 10, 0, 0)
    
    # 1. Mock Trades
    trades = [
        Trade(
            strategy_name="Alpha", trade_id=1, entry_time=t, exit_time=t + timedelta(hours=2),
            side="Long", entry_price=100.0, exit_price=105.0, contracts=10.0, position_value=1000.0,
            commission=2.0, profit=48.0, profit_percent=4.8, holding_time=timedelta(hours=2)
        ),
        Trade(
            strategy_name="Beta", trade_id=1, entry_time=t + timedelta(days=1), exit_time=t + timedelta(days=1, hours=4),
            side="Short", entry_price=50.0, exit_price=52.0, contracts=20.0, position_value=1000.0,
            commission=2.0, profit=-42.0, profit_percent=-4.2, holding_time=timedelta(hours=4)
        ),
        Trade(
            strategy_name="Alpha", trade_id=2, entry_time=t + timedelta(days=2), exit_time=t + timedelta(days=2, hours=1),
            side="Long", entry_price=100.0, exit_price=102.0, contracts=10.0, position_value=1000.0,
            commission=2.0, profit=18.0, profit_percent=1.8, holding_time=timedelta(hours=1)
        ),
    ]

    # 2. Mock Equity Curve
    equity_curve = [
        (t, 10000.0),
        (t + timedelta(hours=2), 10048.0),
        (t + timedelta(days=1), 10048.0),
        (t + timedelta(days=1, hours=4), 10006.0),
        (t + timedelta(days=2), 10006.0),
        (t + timedelta(days=2, hours=1), 10024.0),
    ]

    # 3. Mock Drawdown Curve
    drawdown_curve = [
        (t, 0.0),
        (t + timedelta(hours=2), 0.0),
        (t + timedelta(days=1), 0.0),
        (t + timedelta(days=1, hours=4), 0.42),
        (t + timedelta(days=2), 0.42),
        (t + timedelta(days=2, hours=1), 0.24),
    ]

    # 4. Mock Monthly Returns
    monthly_returns = {
        "2026-01": 0.24,
    }

    # 5. Mock Conflict Logs
    conflict_logs = [
        {
            "conflict_time": t + timedelta(days=1),
            "strategy": "Gamma",
            "trade_id": "Gamma_T_1",
            "required_margin": 500.0,
            "available_margin": 200.0,
            "winner": False,
            "loser": True,
            "skipped_trade": Trade(
                strategy_name="Gamma", trade_id=2, entry_time=t + timedelta(days=1), exit_time=t + timedelta(days=1, hours=2),
                side="Long", entry_price=10.0, exit_price=12.0, contracts=50.0, position_value=500.0,
                commission=1.5, profit=98.5, profit_percent=19.7, holding_time=timedelta(hours=2)
            )
        }
    ]

    # 6. Mock Results Dict
    results = {
        "initial_equity": 10000.0,
        "ending_equity": 10024.0,
        "cagr": 0.03,
        "max_drawdown": 0.42,
        "max_drawdown_cash": 42.0,
        "sharpe": 1.2,
        "sortino": 1.5,
        "calmar": 7.1,
        "ulcer_index": 0.15,
        "recovery_factor": 0.57,
        "equity_curve": equity_curve,
        "drawdown_curve": drawdown_curve,
        "monthly_returns": monthly_returns,
        "trade_statistics": {
            "total_trades": 4,
            "executed_trades": 3,
            "skipped_trades": 1,
            "win_rate": 66.7,
            "net_profit": 24.0,
            "total_profit": 66.0,
            "total_loss": -42.0,
            "profit_factor": 1.57,
            "max_win": 48.0,
            "max_loss": -42.0,
            "avg_trade": 8.0,
            "avg_win": 33.0,
            "avg_loss": -42.0,
            "expectancy": 8.0,
            "max_consecutive_wins": 1,
            "max_consecutive_losses": 1,
        },
        "conflict_frequency": 1,
        "conflict_rate": 0.25,
        "skipped_profit": 98.5,
        "avoided_loss": 0.0,
        "capital_efficiency": 0.15,
        "margin_efficiency": 0.45,
        "time_weighted_capital_efficiency": 0.12,
        "time_weighted_margin_efficiency": 0.41,
        "average_margin_usage": 1500.0,
        "time_weighted_average_margin_usage": 1200.0,
        "peak_margin_usage": 3333.3,
        "average_concurrent_positions": 0.8,
        "maximum_concurrent_positions": 2,
    }

    # 7. Mock Risk Report and Parameters
    risk_report = RiskReport(
        initial_equity=10000.0,
        ending_equity=10024.0,
        peak_equity=10048.0,
        max_drawdown_pct=0.0042,
        total_trades_evaluated=4,
        trades_blocked=1,
        block_reasons={"max_margin_usage": 1},
    )
    
    risk_params = RiskParameters(
        sizing_mode="fixed_capital",
        fixed_capital=1000.0,
        leverage=10.0,
    )

    return results, trades, conflict_logs, risk_report, risk_params


def test_report_suggestions(mock_backtest_data):
    """Verify report generator populates meaningful suggestions."""
    results, trades, conflicts, risk_report, risk_params = mock_backtest_data
    generator = PortfolioReportGenerator(
        backtest_results=results,
        trades=trades,
        conflict_logs=conflicts,
        risk_report=risk_report,
        risk_params=risk_params,
        portfolio_name="Test Report"
    )

    suggestions = generator.generate_suggestions()
    assert len(suggestions) > 0
    # Suggestions should flag skipped profit bottleneck and risk engine blocks
    assert any("Conflict Margin Drag" in s or "skipped profit" in s for s in suggestions)
    assert any("Risk Engine Filter" in s or "blocked" in s for s in suggestions)


def test_pdf_generation(mock_backtest_data):
    """Tests PDF compilation, verifying file existence and valid signature."""
    results, trades, conflicts, risk_report, risk_params = mock_backtest_data
    generator = PortfolioReportGenerator(
        backtest_results=results,
        trades=trades,
        conflict_logs=conflicts,
        risk_report=risk_report,
        risk_params=risk_params,
        portfolio_name="Test Portfolio PDF"
    )

    with tempfile.TemporaryDirectory() as tmpdir:
        pdf_path = os.path.join(tmpdir, "report.pdf")
        generator.generate_pdf(pdf_path)

        assert os.path.exists(pdf_path)
        assert os.path.getsize(pdf_path) > 0

        # Validate PDF signature
        with open(pdf_path, "rb") as f:
            header = f.read(5)
            assert header == b"%PDF-"


def test_excel_generation(mock_backtest_data):
    """Tests Excel compilation, verifying sheets and columns."""
    results, trades, conflicts, risk_report, risk_params = mock_backtest_data
    generator = PortfolioReportGenerator(
        backtest_results=results,
        trades=trades,
        conflict_logs=conflicts,
        risk_report=risk_report,
        risk_params=risk_params,
        portfolio_name="Test Portfolio Excel"
    )

    with tempfile.TemporaryDirectory() as tmpdir:
        excel_path = os.path.join(tmpdir, "report.xlsx")
        generator.generate_excel(excel_path)

        assert os.path.exists(excel_path)
        assert os.path.getsize(excel_path) > 0

        # Open workbook and check sheets
        wb = openpyxl.load_workbook(excel_path)
        expected_sheets = ["Summary Dashboard", "Monthly Matrix", "Executed Trades", "Margin Conflicts"]
        for s in expected_sheets:
            assert s in wb.sheetnames

        # Check Summary Dashboard title
        ws = wb["Summary Dashboard"]
        assert "PORTFOLIO BACKTEST SUMMARY" in ws["A1"].value


def test_csv_generation(mock_backtest_data):
    """Tests CSV export, verifying multiple file segments."""
    results, trades, conflicts, risk_report, risk_params = mock_backtest_data
    generator = PortfolioReportGenerator(
        backtest_results=results,
        trades=trades,
        conflict_logs=conflicts,
        risk_report=risk_report,
        risk_params=risk_params,
        portfolio_name="Test Portfolio CSV"
    )

    with tempfile.TemporaryDirectory() as tmpdir:
        prefix = os.path.join(tmpdir, "test_run")
        files = generator.generate_csv(prefix)

        expected_keys = ["summary", "monthly_returns", "trades", "conflicts"]
        for key in expected_keys:
            assert key in files
            path = files[key]
            assert os.path.exists(path)
            assert os.path.getsize(path) > 0


def test_dynamic_leverage_and_risk_limits(mock_backtest_data):
    """Verify report generator handles dynamic leverage, risk parameters, and halt limits."""
    results, trades, conflicts, risk_report, _ = mock_backtest_data
    # Inject leverage directly into results
    results["leverage"] = 5.0

    # Initialize generator with risk_params=None
    generator = PortfolioReportGenerator(
        backtest_results=results,
        trades=trades,
        conflict_logs=conflicts,
        risk_report=risk_report,
        risk_params=None,
        portfolio_name="Test Dynamic Limits"
    )

    # Compile the flowable objects/story to check they compile without error
    from reportlab.platypus import SimpleDocTemplate
    with tempfile.TemporaryDirectory() as tmpdir:
        pdf_path = os.path.join(tmpdir, "report.pdf")
        generator.generate_pdf(pdf_path)
        assert os.path.exists(pdf_path)
        assert os.path.getsize(pdf_path) > 0
