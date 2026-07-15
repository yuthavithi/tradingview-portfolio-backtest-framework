"""
Portfolio Report Generator Module.

This module compiles backtest results (performance, trades, conflicts, risk metrics)
into a professional multi-page PDF report with custom Matplotlib charts,
an Excel spreadsheet with formatted sheets, and structured CSV files.
"""

import os
import io
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple, Union

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")  # Non-interactive backend for server/command-line generation
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak, KeepTogether
)
from reportlab.pdfgen import canvas

from portfolio.importer import Trade
from portfolio.risk import RiskReport, RiskParameters

# Setup logger for the module
logger = logging.getLogger("portfolio.report")
if not logger.handlers:
    handler = logging.StreamHandler()
    formatter = logging.Formatter(
        "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
    )
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    logger.setLevel(logging.INFO)


class NumberedCanvas(canvas.Canvas):
    """
    Two-pass ReportLab Canvas to dynamically calculate the total page count
    and draw running headers and footers on all pages.
    """

    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self._saved_page_states: List[Dict[str, Any]] = []

    def showPage(self) -> None:
        # Save state of current page for the second rendering pass
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self) -> None:
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count: int) -> None:
        self.saveState()
        
        # Primary colors matching our slate theme
        slate_dark = colors.HexColor("#0f172a") # Slate-900
        slate_medium = colors.HexColor("#475569") # Slate-600
        border_light = colors.HexColor("#e2e8f0") # Slate-200
        
        # Margins bounds: X from 54 to 558 (width 504), Y from 54 to 738 (height 684)
        left_margin = 54
        right_margin = 558
        
        # --- Running Header ---
        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(slate_dark)
        self.drawString(left_margin, 750, "PORTFOLIO BACKTEST PERFORMANCE REPORT")
        
        self.setFont("Helvetica", 8)
        self.setFillColor(slate_medium)
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        self.drawRightString(right_margin, 750, f"Generated: {now_str}")
        
        # Header Rule
        self.setStrokeColor(border_light)
        self.setLineWidth(0.75)
        self.line(left_margin, 742, right_margin, 742)
        
        # --- Running Footer ---
        self.line(left_margin, 52, right_margin, 52)
        
        self.setFont("Helvetica", 8)
        self.setFillColor(slate_medium)
        self.drawString(left_margin, 38, "TradingView Portfolio Backtest Framework")
        
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(right_margin, 38, page_text)
        
        self.restoreState()


class PortfolioReportGenerator:
    """
    Generates PDF, Excel, and CSV performance reports from TradingView backtester outputs.
    """

    def __init__(
        self,
        backtest_results: Dict[str, Any],
        trades: List[Trade],
        conflict_logs: Optional[List[Dict[str, Any]]] = None,
        risk_report: Optional[Union[RiskReport, Dict[str, Any]]] = None,
        risk_params: Optional[Union[RiskParameters, Dict[str, Any]]] = None,
        portfolio_name: str = "TradingView Portfolio",
    ) -> None:
        """
        Initializes the Portfolio Report Generator.

        Args:
            backtest_results: Output dictionary from SharedCapitalEngine.run or similar.
            trades: List of raw/evaluated Trade dataclass objects.
            conflict_logs: Optional list of margin conflict dictionaries.
            risk_report: Optional RiskReport instance or dictionary.
            risk_params: Optional RiskParameters instance or dictionary.
            portfolio_name: Human-readable name of the portfolio.
        """
        self.results = backtest_results
        self.trades = trades
        self.portfolio_name = portfolio_name

        # Extract conflicts if not passed explicitly
        self.conflict_logs = conflict_logs or backtest_results.get("conflict_report") or []

        # Parse risk report and parameters
        self.risk_report = risk_report or backtest_results.get("risk_report")
        self.risk_params = risk_params or backtest_results.get("risk_parameters")

        # Dynamically compute and inject advanced portfolio analytics if missing
        from portfolio.analytics import PortfolioAnalytics
        equity_curve = self.results.get("equity_curve", [])
        initial_equity = self.results.get("initial_equity", 1000.0)
        if equity_curve:
            analytics = PortfolioAnalytics(
                equity_curve=equity_curve,
                trades=self.trades,
                initial_equity=initial_equity,
                trading_days=365
            )
            self.results.setdefault("sharpe", analytics.calculate_sharpe_ratio())
            self.results.setdefault("sortino", analytics.calculate_sortino_ratio())
            self.results.setdefault("calmar", analytics.calculate_calmar_ratio())
            self.results.setdefault("ulcer_index", analytics.calculate_ulcer_index())
            self.results.setdefault("recovery_factor", analytics.calculate_recovery_factor())
            self.results.setdefault("max_drawdown_cash", analytics.calculate_max_drawdown_cash())

        # Configure charts style
        plt.style.use("seaborn-v0_8-whitegrid" if "seaborn-v0_8-whitegrid" in plt.style.available else "default")
        plt.rcParams["font.sans-serif"] = ["Helvetica", "Arial", "DejaVu Sans", "sans-serif"]
        plt.rcParams["font.family"] = "sans-serif"

    # ==========================================================================
    # Suggestions Engine
    # ==========================================================================

    def generate_suggestions(self) -> List[str]:
        """
        Generates dynamic, actionable suggestions based on backtest performance metrics.
        """
        suggestions = []

        # 1. Drawdown Analysis
        max_dd = self.results.get("max_drawdown", 0.0)
        # Convert fraction to percent if small
        if 0.0 < max_dd <= 1.0:
            max_dd *= 100.0

        if max_dd > 20.0:
            suggestions.append(
                f"<b>Risk Sizing Optimization</b>: The backtest recorded a high maximum drawdown of {max_dd:.2f}%. "
                f"Consider lowering the position size multiplier, reducing the risk-per-trade percentage, "
                f"or applying a portfolio-level drawdown stop parameter in the Risk Engine."
            )
        elif 0.0 < max_dd < 5.0 and self.results.get("ending_equity", 0.0) > self.results.get("initial_equity", 0.0):
            suggestions.append(
                f"<b>Leverage Potential</b>: The maximum drawdown is very low ({max_dd:.2f}%). "
                f"There may be room to safely scale up position sizes or increase leverage slightly "
                f"to boost absolute returns, provided historical volatility remains stable."
            )

        # 2. Capital and Margin Efficiency
        margin_eff = self.results.get("margin_efficiency", 0.0)
        if 0.0 < margin_eff <= 1.0:
            margin_eff *= 100.0

        if 0.0 < margin_eff < 40.0 and self.results.get("peak_margin_usage", 0.0) > 0:
            suggestions.append(
                f"<b>Capital Allocation Efficiency</b>: The average margin usage is relatively low compared "
                f"to peak usage (Margin Efficiency: {margin_eff:.2f}%). This indicates that capital remains idle "
                f"for large periods. Consider adding uncorrelated strategies to smooth capital usage, or implementing "
                f"a dynamic capital allocation scheme that reallocates unused funds."
            )

        # 3. Concurrency and Conflict Analysis
        conflict_rate = self.results.get("conflict_rate", 0.0)
        if 0.0 < conflict_rate <= 1.0:
            conflict_rate *= 100.0

        trade_stats = self.results.get("trade_statistics", {})
        skipped_trades = trade_stats.get("skipped_trades", 0)
        skipped_profit = self.results.get("skipped_profit", 0.0)
        avoided_loss = self.results.get("avoided_loss", 0.0)

        if conflict_rate > 10.0:
            if skipped_trades > 0:
                suggestions.append(
                    f"<b>Concurrency Conflict Mitigation</b>: A high conflict rate of {conflict_rate:.2f}% was observed, "
                    f"meaning strategies frequently request margin simultaneously when available capital is insufficient. "
                    f"This led to {skipped_trades} skipped trades. "
                    f"To mitigate this, you could: (1) increase initial capital, (2) run strategies on non-overlapping "
                    f"timeframes, or (3) refine the priority rules so that strategies with higher historical win rates "
                    f"are allocated margin first."
                )
            else:
                suggestions.append(
                    f"<b>Concurrency Overlap</b>: A high overlap rate of {conflict_rate:.2f}% was observed, "
                    f"meaning strategies frequently request margin simultaneously. While current capital was "
                    f"sufficient to execute all trades without skips, future capital constraints or larger size parameters "
                    f"could trigger margin conflicts. Consider this overlap when configuring allocation weights."
                )

        if skipped_profit > avoided_loss:
            net_impact = skipped_profit - avoided_loss
            suggestions.append(
                f"<b>Conflict Margin Drag</b>: The trades skipped due to capital conflicts would have generated a net profit "
                f"of ${net_impact:,.2f} (skipped profit of ${skipped_profit:,.2f} vs avoided losses of ${avoided_loss:,.2f}). "
                f"This represents a significant opportunity cost. Consider expanding your capital base or sizing down trades "
                f"so fewer profitable trades are skipped."
            )
        elif avoided_loss > skipped_profit:
            net_benefit = avoided_loss - skipped_profit
            suggestions.append(
                f"<b>Conflict Margin Protection</b>: The capital conflict filter saved the account a net of "
                f"${net_benefit:,.2f} by blocking trades that would have resulted in losses (avoided loss of "
                f"${avoided_loss:,.2f} vs skipped profit of ${skipped_profit:,.2f}). The margin allocation "
                f"restrictions acted as a protective buffer."
            )

        # 4. Strategy Metrics
        trade_stats = self.results.get("trade_statistics", {})
        pf = trade_stats.get("profit_factor", 0.0)
        wr = trade_stats.get("win_rate", 0.0)
        expectancy = trade_stats.get("expectancy", 0.0)

        if pf != float("inf") and 0.0 < pf < 1.2:
            suggestions.append(
                f"<b>Strategy Profitability Review</b>: The portfolio profit factor is low ({pf:.2f}). "
                f"A profit factor below 1.2 indicates thin margins of profitability. Recommend re-evaluating "
                f"the entry filters or adding stop-loss/take-profit tightening rules to improve the risk/reward ratio."
            )
        if expectancy < 0:
            suggestions.append(
                f"<b>Negative Expectancy Alert</b>: The trade expectancy is negative (${expectancy:,.2f}). "
                f"In the long run, this system is expected to lose money per trade. Re-calibration or disabling "
                f"underperforming component strategies is highly recommended."
            )
        elif wr > 70.0 and pf < 1.5:
            suggestions.append(
                f"<b>High Win Rate but Low Payoff</b>: The system has a high win rate ({wr:.2f}%) but a moderate profit factor "
                f"({pf:.2f}). This suggests a distribution of 'many small wins and occasional large losses'. "
                f"Ensure that trailing stops or hard stops are implemented to avoid catastrophic tail risk."
            )

        # 5. Risk Limits Block Reasons
        if self.risk_report:
            blocked_count = getattr(self.risk_report, "trades_blocked", 0)
            reasons = getattr(self.risk_report, "block_reasons", {})
            if blocked_count > 0:
                reasons_str = ", ".join([f"{k}: {v}" for k, v in reasons.items() if v > 0])
                suggestions.append(
                    f"<b>Risk Engine Filter</b>: The Risk Engine blocked {blocked_count} trades due to limit breaches "
                    f"({reasons_str}). This prevented excessive exposure. Check if limits are too tight for standard "
                    f"market regimes."
                )

        # Default suggestion
        if not suggestions:
            suggestions.append(
                "<b>Portfolio Status Healthy</b>: The portfolio backtest shows positive metrics and balanced risk. "
                "Maintain regular monitoring of strategy correlation to avoid simultaneous drawdown overlaps."
            )

        return suggestions

    # ==========================================================================
    # Matplotlib Charts Generator
    # ==========================================================================

    def _generate_equity_drawdown_plot(self) -> io.BytesIO:
        """
        Generates a combined Equity and Drawdown curve plot as a byte buffer.
        """
        eq_curve = self.results.get("equity_curve", [])
        dd_curve = self.results.get("drawdown_curve", [])

        if not eq_curve:
            # Fallback
            fig, ax = plt.subplots(figsize=(8, 4))
            ax.text(0.5, 0.5, "No Curve Data", ha="center", va="center")
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=200, bbox_inches="tight")
            plt.close(fig)
            buf.seek(0)
            return buf

        df_eq = pd.DataFrame(eq_curve, columns=["timestamp", "equity"])
        df_eq["timestamp"] = pd.to_datetime(df_eq["timestamp"])
        df_eq = df_eq.set_index("timestamp")

        # Set up a multi-plot layout (2 subplots stacked)
        fig, (ax_eq, ax_dd) = plt.subplots(2, 1, figsize=(8, 5.5), sharex=True,
                                           gridspec_kw={"height_ratios": [2.5, 1.2]})

        # Color palette
        color_navy = "#1e293b"
        color_emerald = "#059669"
        color_rose = "#dc2626"

        # 1. Equity Plot
        ax_eq.plot(df_eq.index, df_eq["equity"], color=color_navy, linewidth=2, label="Equity Curve")
        ax_eq.fill_between(df_eq.index, df_eq["equity"], self.results.get("initial_equity", 1000.0),
                           color="#cbd5e1", alpha=0.3)
        ax_eq.set_title("Portfolio Equity & Drawdown Curves", fontsize=12, fontweight="bold", pad=8)
        ax_eq.set_ylabel("Account Value ($)", fontsize=9)
        ax_eq.grid(True, linestyle="--", alpha=0.5)
        ax_eq.tick_params(labelsize=8)

        # 2. Drawdown Plot
        if dd_curve:
            df_dd = pd.DataFrame(dd_curve, columns=["timestamp", "drawdown"])
            df_dd["timestamp"] = pd.to_datetime(df_dd["timestamp"])
            df_dd = df_dd.set_index("timestamp")
        else:
            # Calculate drawdown
            peaks = df_eq["equity"].cummax()
            df_dd = pd.DataFrame((df_eq["equity"] - peaks) / peaks * 100.0, columns=["drawdown"])

        ax_dd.plot(df_dd.index, -df_dd["drawdown"], color=color_rose, linewidth=1.2, label="Drawdown %")
        ax_dd.fill_between(df_dd.index, -df_dd["drawdown"], 0, color=color_rose, alpha=0.15)
        ax_dd.set_ylabel("Drawdown (%)", fontsize=9)
        ax_dd.set_xlabel("Date", fontsize=9)
        ax_dd.grid(True, linestyle="--", alpha=0.5)
        ax_dd.tick_params(labelsize=8)

        # Format dates nicely
        fig.autofmt_xdate(rotation=15, ha="right")

        plt.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=200, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        return buf

    def _generate_monthly_heatmap_plot(self) -> io.BytesIO:
        """
        Generates a Monthly Return Heatmap plot as a byte buffer.
        """
        monthly_returns = self.results.get("monthly_returns", {})
        if not monthly_returns:
            fig, ax = plt.subplots(figsize=(6, 3))
            ax.text(0.5, 0.5, "No Monthly Return Data", ha="center", va="center")
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=200, bbox_inches="tight")
            plt.close(fig)
            buf.seek(0)
            return buf

        # Extract monthly data into a structured frame
        rows = []
        for key, val in monthly_returns.items():
            # key can be "2026-01" or (2026, 1)
            if isinstance(key, str):
                parts = key.split("-")
                year, month = int(parts[0]), int(parts[1])
            else:
                year, month = key[0], key[1]
            rows.append({"Year": year, "Month": month, "Return": val})

        df = pd.DataFrame(rows)
        pivot_df = df.pivot(index="Year", columns="Month", values="Return")

        # Fill missing months
        for m in range(1, 13):
            if m not in pivot_df.columns:
                pivot_df[m] = np.nan
        pivot_df = pivot_df.reindex(columns=range(1, 13))
        pivot_df.columns = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

        fig, ax = plt.subplots(figsize=(8, 3))

        # We can implement a clean custom heatmap using matplotlib's imshow or pcolor
        # Mask NaNs so they display differently (e.g., light grey/white background)
        data = pivot_df.values
        masked_data = np.ma.masked_invalid(data)

        # Dynamic colormap diverging from red to green
        cmap = plt.cm.RdYlGn
        cmap.set_bad(color="#f1f5f9")  # slate-100

        im = ax.imshow(masked_data, cmap=cmap, aspect="auto", interpolation="nearest",
                       vmin=-10, vmax=10) # Bounded for visual contrast

        # Disable grid lines overlaying the cells
        ax.grid(False)

        # Show ticks and labels
        ax.set_xticks(np.arange(12))
        ax.set_xticklabels(pivot_df.columns, fontsize=8)
        ax.set_yticks(np.arange(len(pivot_df.index)))
        ax.set_yticklabels(pivot_df.index.astype(str), fontsize=8, fontweight="bold")
        ax.set_title("Monthly Return Breakdown (%)", fontsize=11, fontweight="bold", pad=8)

        # Turn spines off
        for edge in ["top", "bottom", "left", "right"]:
            ax.spines[edge].set_visible(False)

        # Annotate cell values
        for i in range(len(pivot_df.index)):
            for j in range(12):
                val = data[i, j]
                if not np.isnan(val):
                    text_color = "black" if abs(val) < 6 else "white"
                    ax.text(j, i, f"{val:+.1f}%", ha="center", va="center", color=text_color,
                            fontsize=7.5, fontweight="bold")

        plt.colorbar(im, ax=ax, shrink=0.7, aspect=15, pad=0.03).ax.tick_params(labelsize=8)
        plt.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=200, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        return buf

    def _generate_yearly_returns_plot(self) -> io.BytesIO:
        """
        Generates a Yearly Return Bar Chart as a byte buffer.
        """
        monthly_returns = self.results.get("monthly_returns", {})
        if not monthly_returns:
            fig, ax = plt.subplots(figsize=(6, 2.5))
            ax.text(0.5, 0.5, "No Yearly Data", ha="center", va="center")
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=200, bbox_inches="tight")
            plt.close(fig)
            buf.seek(0)
            return buf

        # Group by Year
        yearly_pnl = {}
        for key, val in monthly_returns.items():
            if isinstance(key, str):
                year = int(key.split("-")[0])
            else:
                year = key[0]
            # Sum returns or compound them. Summing monthly percentage returns is a standard approximation.
            yearly_pnl[year] = yearly_pnl.get(year, 0.0) + val

        years = sorted(list(yearly_pnl.keys()))
        returns = [yearly_pnl[y] for y in years]

        fig, ax = plt.subplots(figsize=(7, 2.8))

        colors = ["#10b981" if r >= 0 else "#ef4444" for r in returns]
        bars = ax.bar([str(y) for y in years], returns, color=colors, edgecolor="none", width=0.45)
        ax.axhline(0, color="#475569", linewidth=0.75, linestyle="--")

        ax.set_title("Year-Over-Year Performance (%)", fontsize=11, fontweight="bold", pad=8)
        ax.set_ylabel("Net Return (%)", fontsize=8.5)
        ax.tick_params(labelsize=8.5)
        ax.grid(True, axis="y", linestyle="--", alpha=0.5)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

        # Label values on top of bars
        for bar in bars:
            height = bar.get_height()
            label_y = height + 0.8 if height >= 0 else height - 1.8
            ax.annotate(
                f"{height:+.1f}%",
                xy=(bar.get_x() + bar.get_width() / 2, height),
                xytext=(0, 3 if height >= 0 else -10),
                textcoords="offset points",
                ha="center", va="bottom", fontsize=8, fontweight="bold"
            )

        plt.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=200, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        return buf

    def _generate_capital_usage_plot(self) -> io.BytesIO:
        """
        Generates a chart of capital and margin usage over time.
        """
        history = self.results.get("equity_curve", [])
        if not history:
            fig, ax = plt.subplots(figsize=(8, 3))
            ax.text(0.5, 0.5, "No History Data", ha="center", va="center")
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=200, bbox_inches="tight")
            plt.close(fig)
            buf.seek(0)
            return buf

        # Generate custom mock margin if history doesn't have it (for backward compatibility)
        timestamps = [h[0] for h in history]
        equities = [h[1] for r, h in enumerate(history)]

        fig, ax = plt.subplots(figsize=(8, 3.2))

        # Plot equity curve
        ax.plot(timestamps, equities, color="#0f172a", label="Total Equity", linewidth=1.5)

        # Plot synthetic/actual margin usage
        # Check if we have self.results.get("history") containing raw snapshots with margin_used
        raw_history = self.results.get("history", [])
        if raw_history:
            hist_df = pd.DataFrame(raw_history)
            hist_df["timestamp"] = pd.to_datetime(hist_df["timestamp"])
            ax.fill_between(hist_df["timestamp"], hist_df["margin_used"], 0,
                            color="#3b82f6", alpha=0.25, label="Allocated Margin")
            ax.plot(hist_df["timestamp"], hist_df["margin_used"], color="#3b82f6", linewidth=1.0, alpha=0.8)
        else:
            # Generate a baseline for visualization
            init_eq = self.results.get("initial_equity", 1000.0)
            avg_margin = self.results.get("average_margin_usage", init_eq * 0.2)
            ax.axhline(avg_margin, color="#3b82f6", linestyle="--", linewidth=1, label="Avg Margin Usage")

        ax.set_title("Margin Allocation vs. Account Equity", fontsize=11, fontweight="bold", pad=8)
        ax.set_ylabel("Capital / Margin ($)", fontsize=8.5)
        ax.tick_params(labelsize=8)
        ax.grid(True, linestyle="--", alpha=0.5)
        ax.legend(loc="upper left", fontsize=8)
        fig.autofmt_xdate(rotation=15, ha="right")

        plt.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=200, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        return buf

    # ==========================================================================
    # PDF Report Compiler
    # ==========================================================================

    def generate_pdf(self, output_path: str) -> None:
        """
        Compiles the professional PDF performance report.

        Args:
            output_path: Destination file path for the PDF report.
        """
        logger.info(f"Generating PDF report: {output_path}")

        # Page layout configuration
        # letter size is 612 x 792 pt
        # Margins left/right = 54 pt, top/bottom = 72 pt
        doc = SimpleDocTemplate(
            output_path,
            pagesize=letter,
            leftMargin=54,
            rightMargin=54,
            topMargin=72,
            bottomMargin=72
        )

        styles = getSampleStyleSheet()
        
        # Define clean, professional Paragraph Styles matching Slate theme
        style_title = ParagraphStyle(
            "DocTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=22,
            leading=26,
            textColor=colors.HexColor("#0f172a"),
            alignment=0, # Left-aligned
            spaceAfter=6
        )

        style_subtitle = ParagraphStyle(
            "DocSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#475569"),
            spaceAfter=20
        )

        style_h1 = ParagraphStyle(
            "SectionHeader",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#0f172a"),
            spaceBefore=14,
            spaceAfter=8,
            keepWithNext=True
        )

        style_body = ParagraphStyle(
            "BodyText",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9.5,
            leading=13,
            textColor=colors.HexColor("#334155"),
            spaceAfter=8
        )

        style_suggestion = ParagraphStyle(
            "SuggestionBullet",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=12.5,
            textColor=colors.HexColor("#1e293b"),
            leftIndent=15,
            firstLineIndent=-10,
            spaceAfter=6
        )

        style_table_header = ParagraphStyle(
            "TableHeader",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8.5,
            leading=11,
            textColor=colors.white,
            alignment=0
        )

        style_table_cell = ParagraphStyle(
            "TableCell",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            textColor=colors.HexColor("#334155"),
            alignment=0
        )

        style_table_cell_bold = ParagraphStyle(
            "TableCellBold",
            parent=style_table_cell,
            fontName="Helvetica-Bold"
        )

        story: List[Any] = []

        # --- Document Header Block ---
        story.append(Paragraph(self.portfolio_name, style_title))
        story.append(Paragraph(f"Performance Backtesting Analysis Report | Issued {datetime.now().strftime('%Y-%m-%d')}",
                               style_subtitle))

        # --- Section 1: Executive Summary ---
        story.append(Paragraph("Executive Summary", style_h1))
        
        # Calculate summary parameters
        initial_equity = self.results.get("initial_equity", 1000.0)
        ending_equity = self.results.get("ending_equity", 1000.0)
        net_profit = ending_equity - initial_equity
        net_profit_pct = (net_profit / initial_equity) * 100.0 if initial_equity > 0 else 0.0
        
        start_date = "N/A"
        end_date = "N/A"
        if self.trades:
            sorted_trades = sorted(self.trades, key=lambda t: t.entry_time)
            start_date = sorted_trades[0].entry_time.strftime("%Y-%m-%d")
            end_date = sorted_trades[-1].exit_time.strftime("%Y-%m-%d") if sorted_trades[-1].exit_time else sorted_trades[-1].entry_time.strftime("%Y-%m-%d")

        summary_para = (
            f"This backtest report summarizes the quantitative simulation of the portfolio <b>{self.portfolio_name}</b> "
            f"covering the period from <b>{start_date}</b> to <b>{end_date}</b>. "
            f"Starting with an initial capital of <b>${initial_equity:,.2f}</b>, the portfolio concluded the simulation "
            f"with an ending equity of <b>${ending_equity:,.2f}</b>, generating a net PnL of <b>${net_profit:+,.2f}</b> "
            f"(<b>{net_profit_pct:+.2f}%</b>). A total of <b>{len(self.trades)}</b> trades were evaluated across all strategies. "
            f"The simulation enforced margin leverage constraint checks, position sizing parameters, and risk limits as defined in the rules."
        )
        story.append(Paragraph(summary_para, style_body))
        story.append(Spacer(1, 10))

        # --- Section 2: Portfolio Performance ---
        story.append(Paragraph("Portfolio Performance & Risk Ratios", style_h1))
        
        cagr = self.results.get("cagr", 0.0) * 100.0
        max_dd = self.results.get("max_drawdown", 0.0)
            
        # Draw table of performance statistics
        perf_data = [
            [Paragraph("Performance Metric", style_table_header), Paragraph("Value", style_table_header),
             Paragraph("Risk Ratio Metric", style_table_header), Paragraph("Value", style_table_header)],
            [Paragraph("Initial Capital", style_table_cell), Paragraph(f"${initial_equity:,.2f}", style_table_cell_bold),
             Paragraph("Sharpe Ratio", style_table_cell), Paragraph(f"{self.results.get('sharpe', 0.0):.2f}", style_table_cell_bold)],
            [Paragraph("Ending Capital", style_table_cell), Paragraph(f"${ending_equity:,.2f}", style_table_cell_bold),
             Paragraph("Sortino Ratio", style_table_cell), Paragraph(f"{self.results.get('sortino', 0.0):.2f}", style_table_cell_bold)],
            [Paragraph("Net Profit ($)", style_table_cell), Paragraph(f"${net_profit:+,.2f}", style_table_cell_bold),
             Paragraph("Calmar / MAR Ratio", style_table_cell), Paragraph(f"{self.results.get('calmar', self.results.get('mar', 0.0)):.2f}", style_table_cell_bold)],
            [Paragraph("Net Profit (%)", style_table_cell), Paragraph(f"{net_profit_pct:+.2f}%", style_table_cell_bold),
             Paragraph("Ulcer Index", style_table_cell), Paragraph(f"{self.results.get('ulcer_index', 0.0):.2f}", style_table_cell_bold)],
            [Paragraph("CAGR", style_table_cell), Paragraph(f"{cagr:.2f}%", style_table_cell_bold),
             Paragraph("Recovery Factor", style_table_cell), Paragraph(f"{self.results.get('recovery_factor', 0.0):.2f}", style_table_cell_bold)],
            [Paragraph("Max Drawdown (%)", style_table_cell), Paragraph(f"{max_dd:.2f}%", style_table_cell_bold),
             Paragraph("Max Drawdown ($)", style_table_cell), Paragraph(f"${self.results.get('max_drawdown_cash', 0.0):,.2f}", style_table_cell_bold)]
        ]

        t_perf = Table(perf_data, colWidths=[150, 102, 150, 102])
        t_perf.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")), # Teal header
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
            ("TOPPADDING", (0, 0), (-1, 0), 5),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ]))
        story.append(t_perf)
        story.append(Spacer(1, 15))

        # --- Section 3: Equity & Drawdown Curves ---
        story.append(Paragraph("Equity & Drawdown Curves", style_h1))
        # Embed Matplotlib generated charts
        eq_img_buf = self._generate_equity_drawdown_plot()
        story.append(Image(eq_img_buf, width=480, height=330))
        story.append(Spacer(1, 10))

        story.append(PageBreak())  # Move monthly breakdowns to next page

        # --- Section 4: Monthly Return Matrix ---
        story.append(Paragraph("Monthly Returns Matrix", style_h1))
        
        monthly_returns = self.results.get("monthly_returns", {})
        if monthly_returns:
            # Pivot monthly returns for tabular view
            rows = []
            for key, val in monthly_returns.items():
                if isinstance(key, str):
                    parts = key.split("-")
                    year, month = int(parts[0]), int(parts[1])
                else:
                    year, month = key[0], key[1]
                rows.append({"Year": year, "Month": month, "Return": val})
            df = pd.DataFrame(rows)
            pivot_df = df.pivot(index="Year", columns="Month", values="Return")
            
            # Fill columns 1-12
            for m in range(1, 13):
                if m not in pivot_df.columns:
                    pivot_df[m] = np.nan
            pivot_df = pivot_df.reindex(columns=range(1, 13))
            
            # Construct data table
            headers = ["Year", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "YTD"]
            matrix_data = [[Paragraph(h, style_table_header) for h in headers]]
            
            for yr in pivot_df.index:
                row_cells = [Paragraph(str(yr), style_table_cell_bold)]
                ytd_sum = 0.0
                for m in range(1, 13):
                    val = pivot_df.loc[yr, m]
                    if np.isnan(val):
                        row_cells.append(Paragraph("-", style_table_cell))
                    else:
                        ytd_sum += val
                        color_cell = "#15803d" if val >= 0 else "#be123c"
                        row_cells.append(Paragraph(f"<font color='{color_cell}'>{val:+.1f}%</font>", style_table_cell))
                
                # YTD Cell
                color_ytd = "#15803d" if ytd_sum >= 0 else "#be123c"
                row_cells.append(Paragraph(f"<font color='{color_ytd}'><b>{ytd_sum:+.1f}%</b></font>", style_table_cell))
                matrix_data.append(row_cells)
                
            # Render Table
            # Total width must be 504 -> Year: 48, Months: 35 each, YTD: 36
            col_w = [48] + [35] * 12 + [36]
            t_matrix = Table(matrix_data, colWidths=col_w)
            t_matrix.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(t_matrix)
            story.append(Spacer(1, 10))

        # Embed monthly returns heatmap
        hm_img_buf = self._generate_monthly_heatmap_plot()
        story.append(Image(hm_img_buf, width=480, height=180))
        story.append(Spacer(1, 15))

        # --- Section 5: Annual Returns ---
        story.append(Paragraph("Annual Performance & Yearly Returns", style_h1))
        yr_img_buf = self._generate_yearly_returns_plot()
        story.append(Image(yr_img_buf, width=480, height=180))
        story.append(Spacer(1, 10))

        story.append(PageBreak())

        # --- Section 6: Trade Statistics ---
        story.append(Paragraph("Trade Performance Statistics", style_h1))
        t_stats = self.results.get("trade_statistics", {})
        
        trade_data = [
            [Paragraph("Trade Statistic", style_table_header), Paragraph("Value", style_table_header),
             Paragraph("Trade Statistic", style_table_header), Paragraph("Value", style_table_header)],
            [Paragraph("Total Trades Evaluated", style_table_cell), Paragraph(f"{t_stats.get('total_trades', len(self.trades))}", style_table_cell_bold),
             Paragraph("Average Trade PnL", style_table_cell), Paragraph(f"${t_stats.get('avg_trade', 0.0):+,.2f}", style_table_cell_bold)],
            [Paragraph("Executed Trades Count", style_table_cell), Paragraph(f"{t_stats.get('executed_trades', len(self.trades))}", style_table_cell_bold),
             Paragraph("Average Winning Trade", style_table_cell), Paragraph(f"${t_stats.get('avg_win', 0.0):,.2f}", style_table_cell_bold)],
            [Paragraph("Skipped Trades (Conflicts)", style_table_cell), Paragraph(f"{t_stats.get('skipped_trades', 0)}", style_table_cell_bold),
             Paragraph("Average Losing Trade", style_table_cell), Paragraph(f"${t_stats.get('avg_loss', 0.0):,.2f}", style_table_cell_bold)],
            [Paragraph("Portfolio Win Rate (%)", style_table_cell), Paragraph(f"{t_stats.get('win_rate', 0.0):.1f}%", style_table_cell_bold),
             Paragraph("Largest Winning Trade", style_table_cell), Paragraph(f"${t_stats.get('max_win', 0.0):,.2f}", style_table_cell_bold)],
            [Paragraph("Profit Factor", style_table_cell), Paragraph(f"{t_stats.get('profit_factor', 0.0):.2f}", style_table_cell_bold),
             Paragraph("Largest Losing Trade", style_table_cell), Paragraph(f"${t_stats.get('max_loss', 0.0):,.2f}", style_table_cell_bold)],
            [Paragraph("Profit Expectancy", style_table_cell), Paragraph(f"${t_stats.get('expectancy', 0.0):+,.2f}", style_table_cell_bold),
             Paragraph("Consecutive Streak (W/L)", style_table_cell), Paragraph(f"{t_stats.get('max_consecutive_wins', 0)} W / {t_stats.get('max_consecutive_losses', 0)} L", style_table_cell_bold)]
        ]

        t_trade = Table(trade_data, colWidths=[150, 102, 150, 102])
        t_trade.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4.5),
        ]))
        story.append(t_trade)
        story.append(Spacer(1, 15))

        # --- Section 7: Capital & Margin Usage ---
        story.append(Paragraph("Capital Allocation & Margin Usage", style_h1))
        
        cap_eff = self.results.get("capital_efficiency", 0.0)
        margin_eff = self.results.get("margin_efficiency", 0.0)
        tw_cap_eff = self.results.get("time_weighted_capital_efficiency", 0.0)
        tw_margin_eff = self.results.get("time_weighted_margin_efficiency", 0.0)
        
        # Convert fractions to percent
        if 0.0 < cap_eff <= 1.0: cap_eff *= 100.0
        if 0.0 < margin_eff <= 1.0: margin_eff *= 100.0
        if 0.0 < tw_cap_eff <= 1.0: tw_cap_eff *= 100.0
        if 0.0 < tw_margin_eff <= 1.0: tw_margin_eff *= 100.0
        
        usage_data = [
            [Paragraph("Capital/Margin Metric", style_table_header), Paragraph("Value", style_table_header),
             Paragraph("Capital/Margin Metric", style_table_header), Paragraph("Value", style_table_header)],
            [Paragraph("Peak Margin Requirement", style_table_cell), Paragraph(f"${self.results.get('peak_margin_usage', 0.0):,.2f}", style_table_cell_bold),
             Paragraph("Capital Efficiency (Avg)", style_table_cell), Paragraph(f"{cap_eff:.2f}%", style_table_cell_bold)],
            [Paragraph("Average Margin Usage", style_table_cell), Paragraph(f"${self.results.get('average_margin_usage', 0.0):,.2f}", style_table_cell_bold),
             Paragraph("Time-Weighted Capital Eff.", style_table_cell), Paragraph(f"{tw_cap_eff:.2f}%", style_table_cell_bold)],
            [Paragraph("Time-Weighted Avg Margin", style_table_cell), Paragraph(f"${self.results.get('time_weighted_average_margin_usage', 0.0):,.2f}", style_table_cell_bold),
             Paragraph("Margin Efficiency (Avg)", style_table_cell), Paragraph(f"{margin_eff:.2f}%", style_table_cell_bold)],
            [Paragraph("Max Concurrent Positions", style_table_cell), Paragraph(f"{self.results.get('maximum_concurrent_positions', 0)}", style_table_cell_bold),
             Paragraph("Time-Weighted Margin Eff.", style_table_cell), Paragraph(f"{tw_margin_eff:.2f}%", style_table_cell_bold)],
            [Paragraph("Average Concurrent Positions", style_table_cell), Paragraph(f"{self.results.get('average_concurrent_positions', 0.0):.2f}", style_table_cell_bold),
             Paragraph("Available Account Margin", style_table_cell), Paragraph(f"${self.results.get('available_margin', ending_equity):,.2f}", style_table_cell_bold)]
        ]

        t_usage = Table(usage_data, colWidths=[150, 102, 150, 102])
        t_usage.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4.5),
        ]))
        story.append(t_usage)
        story.append(Spacer(1, 10))

        # Embed capital usage visualization
        cap_img_buf = self._generate_capital_usage_plot()
        story.append(Image(cap_img_buf, width=480, height=180))
        story.append(Spacer(1, 10))

        story.append(PageBreak())

        # --- Section 8: Conflict Analysis ---
        story.append(Paragraph("Capital Conflict Analysis", style_h1))
        
        skipped_profit = self.results.get("skipped_profit", 0.0)
        avoided_loss = self.results.get("avoided_loss", 0.0)
        
        conflict_rate_pct = self.results.get("conflict_rate", 0.0)
        if 0.0 < conflict_rate_pct <= 1.0:
            conflict_rate_pct *= 100.0
            
        conflict_summary_data = [
            [Paragraph("Conflict Metric", style_table_header), Paragraph("Value", style_table_header),
             Paragraph("Conflict Metric", style_table_header), Paragraph("Value", style_table_header)],
            [Paragraph("Conflict Frequency (Events)", style_table_cell), Paragraph(f"{self.results.get('conflict_frequency', 0)}", style_table_cell_bold),
             Paragraph("Skipped Profit (Opportunity Cost)", style_table_cell), Paragraph(f"${skipped_profit:,.2f}", style_table_cell_bold)],
            [Paragraph("Conflict Rate (%)", style_table_cell), Paragraph(f"{conflict_rate_pct:.2f}%", style_table_cell_bold),
             Paragraph("Avoided Loss (Capital Protection)", style_table_cell), Paragraph(f"${avoided_loss:,.2f}", style_table_cell_bold)]
        ]
        
        t_conflict_sum = Table(conflict_summary_data, colWidths=[150, 102, 150, 102])
        t_conflict_sum.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4.5),
        ]))
        story.append(t_conflict_sum)
        story.append(Spacer(1, 10))

        # Recent Conflict Logs Table
        story.append(Paragraph("Margin Allocation Conflict Log Summary", ParagraphStyle("SubHeader", parent=style_h1, fontSize=10, keepWithNext=True)))
        if self.conflict_logs:
            # Table headers: Time, Strategy, Required, Available, Winner/Loser, Skipped Trade PnL
            log_headers = ["Time", "Strategy Name", "Req Margin", "Avail Margin", "Status", "Skipped PnL"]
            log_rows = [[Paragraph(h, style_table_header) for h in log_headers]]
            
            # Show up to 8 conflict logs for display brevity
            display_logs = self.conflict_logs[:8]
            for log in display_logs:
                dt_str = log["conflict_time"]
                if isinstance(dt_str, datetime):
                    dt_str = dt_str.strftime("%m-%d %H:%M")
                    
                status_text = "Winner" if log.get("winner") else "Loser (Skipped)"
                color_status = "#15803d" if log.get("winner") else "#be123c"
                
                skipped_pnl_str = "-"
                if log.get("loser") and log.get("skipped_trade") is not None:
                    trade_obj = log["skipped_trade"]
                    prof = getattr(trade_obj, "profit", 0.0)
                    color_p = "#15803d" if prof >= 0 else "#be123c"
                    skipped_pnl_str = f"<font color='{color_p}'>${prof:+,.2f}</font>"

                log_rows.append([
                    Paragraph(str(dt_str), style_table_cell),
                    Paragraph(log.get("strategy", "Unknown"), style_table_cell),
                    Paragraph(f"${log.get('required_margin', 0.0):,.1f}", style_table_cell),
                    Paragraph(f"${log.get('available_margin', 0.0):,.1f}", style_table_cell),
                    Paragraph(f"<font color='{color_status}'><b>{status_text}</b></font>", style_table_cell),
                    Paragraph(skipped_pnl_str, style_table_cell),
                ])
                
            t_log = Table(log_rows, colWidths=[80, 100, 75, 75, 94, 80])
            t_log.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("TOPPADDING", (0, 0), (-1, -1), 3.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
            ]))
            story.append(t_log)
        else:
            story.append(Paragraph("No margin allocation conflicts were recorded during this simulation run.", style_body))
        story.append(Spacer(1, 15))

        # --- Section 9: Risk Engine Metrics ---
        story.append(Paragraph("Risk Limits & Constraints Metrics", style_h1))
        
        # Sizing model details
        sizing_mode = "N/A"
        leverage_val = 1.0
        max_allowed_risk = "N/A"
        max_daily_loss = "N/A"

        if self.risk_params:
            if isinstance(self.risk_params, dict):
                sizing_mode = self.risk_params.get("sizing_mode", "N/A")
                leverage_val = self.risk_params.get("leverage", 1.0)
                r_pct = self.risk_params.get("max_portfolio_risk_pct")
                d_pct = self.risk_params.get("max_daily_loss_pct")
            else:
                sizing_mode = getattr(self.risk_params, "sizing_mode", "N/A")
                leverage_val = getattr(self.risk_params, "leverage", 1.0)
                r_pct = getattr(self.risk_params, "max_portfolio_risk_pct", None)
                d_pct = getattr(self.risk_params, "max_daily_loss_pct", None)

            if r_pct is not None:
                max_allowed_risk = f"{r_pct * 100:.2f}%"
            if d_pct is not None:
                max_daily_loss = f"{d_pct * 100:.2f}%"
        else:
            # Fall back to simulation leverage if risk parameters are not available
            leverage_val = self.results.get("leverage", 1.0)

        leverage = f"{leverage_val:.1f}x"

        # Risk report metrics
        total_eval = len(self.trades)
        blocked = 0
        reasons_list = []
        halt_breached = "NO"

        if self.risk_report:
            if isinstance(self.risk_report, dict):
                total_eval = self.risk_report.get("total_trades_evaluated", len(self.trades))
                blocked = self.risk_report.get("trades_blocked", 0)
                reasons = self.risk_report.get("block_reasons", {})
                reasons_list = [f"{k}: {v}" for k, v in reasons.items() if v > 0]
                is_halted = self.risk_report.get("trading_halted", False)
            else:
                total_eval = getattr(self.risk_report, "total_trades_evaluated", len(self.trades))
                blocked = getattr(self.risk_report, "trades_blocked", 0)
                reasons = getattr(self.risk_report, "block_reasons", {})
                reasons_list = [f"{k}: {v}" for k, v in reasons.items() if v > 0]
                is_halted = getattr(self.risk_report, "trading_halted", False)

            if is_halted:
                halt_breached = "YES"

        reasons_text = ", ".join(reasons_list) if reasons_list else "None"

        risk_data = [
            [Paragraph("Risk Parameter / Constraint", style_table_header), Paragraph("Value", style_table_header),
             Paragraph("Risk Observation / Limit Metrics", style_table_header), Paragraph("Value", style_table_header)],
            [Paragraph("Position Sizing Model", style_table_cell), Paragraph(str(sizing_mode).upper(), style_table_cell_bold),
             Paragraph("Total Trades Evaluated", style_table_cell), Paragraph(str(total_eval), style_table_cell_bold)],
            [Paragraph("Leverage Applied", style_table_cell), Paragraph(str(leverage), style_table_cell_bold),
             Paragraph("Trades Blocked by Risk", style_table_cell), Paragraph(str(blocked), style_table_cell_bold)],
            [Paragraph("Max Allowed Portfolio Risk", style_table_cell), Paragraph(max_allowed_risk, style_table_cell_bold),
             Paragraph("Halt Limit Breached", style_table_cell), Paragraph(halt_breached, style_table_cell_bold)],
            [Paragraph("Max Daily Loss Limit", style_table_cell), Paragraph(max_daily_loss, style_table_cell_bold),
             Paragraph("Risk Block Reasons Detail", style_table_cell), Paragraph(reasons_text, style_table_cell_bold)]
        ]

        t_risk = Table(risk_data, colWidths=[150, 102, 150, 102])
        t_risk.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4.5),
        ]))
        story.append(t_risk)
        story.append(Spacer(1, 15))

        # --- Section 10: Monte Carlo Stress Test & CVaR Analysis ---
        cvar_val = self.results.get("cvar")
        if cvar_val is not None:
            story.append(Paragraph("Monte Carlo Stress Test & CVaR Analysis", style_h1))
            
            cvar_pct = cvar_val * 100.0
            conf_level = self.results.get("cvar_confidence", 0.95) * 100.0
            mc_paths = self.results.get("cvar_iterations", 1000)
            opt_obj = self.results.get("optimization_objective", "None (Baseline)").upper()
            
            # Format weights list
            opt_w = self.results.get("optimal_weights")
            if opt_w:
                w_str = ", ".join([f"{k}: {v*100:.1f}%" for k, v in opt_w.items()])
            else:
                w_str = "Equal Weight (100% per strategy)"
                
            cvar_data = [
                [Paragraph("Stress Test Metric", style_table_header), Paragraph("Value", style_table_header),
                 Paragraph("Simulation Parameter", style_table_header), Paragraph("Value", style_table_header)],
                [Paragraph("Conditional Drawdown at Risk (CVaR)", style_table_cell), Paragraph(f"<b>{cvar_pct:.2f}%</b>", style_table_cell_bold),
                 Paragraph("Monte Carlo Bootstrap Paths", style_table_cell), Paragraph(str(mc_paths), style_table_cell_bold)],
                [Paragraph("Confidence Level", style_table_cell), Paragraph(f"{conf_level:.1f}%", style_table_cell_bold),
                 Paragraph("Optimization Objective", style_table_cell), Paragraph(opt_obj, style_table_cell_bold)],
                [Paragraph("Stress Test Floating PnL Mode", style_table_cell), Paragraph("MAE (Worst-Case Adverse Excursion)" if self.results.get("stress_test_drawdown") else "Linear Interpolation", style_table_cell_bold),
                 Paragraph("Allocation Mode", style_table_cell), Paragraph("Optimized Weights" if opt_w else "Equal Weight (100%)", style_table_cell_bold)]
            ]
            
            t_cvar = Table(cvar_data, colWidths=[150, 102, 150, 102])
            t_cvar.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4.5),
            ]))
            story.append(t_cvar)
            story.append(Spacer(1, 10))

            # Dedicated Weights Table
            style_h2 = ParagraphStyle("SubHeader", parent=style_h1, fontSize=11, leading=14, spaceBefore=6, spaceAfter=4, keepWithNext=True)
            if opt_w:
                story.append(Paragraph("Optimized Strategy Allocation Weights", style_h2))
                
                # Split weights into 4-column side-by-side format
                items = list(opt_w.items())
                half = (len(items) + 1) // 2
                
                weights_data = [
                    [Paragraph("Strategy (Symbol)", style_table_header), Paragraph("Weight", style_table_header),
                     Paragraph("Strategy (Symbol)", style_table_header), Paragraph("Weight", style_table_header)]
                ]
                
                for idx in range(half):
                    k1, v1 = items[idx]
                    col1 = Paragraph(k1, style_table_cell)
                    col2 = Paragraph(f"<b>{v1*100:.2f}%</b>", style_table_cell_bold)
                    
                    if idx + half < len(items):
                        k2, v2 = items[idx + half]
                        col3 = Paragraph(k2, style_table_cell)
                        col4 = Paragraph(f"<b>{v2*100:.2f}%</b>", style_table_cell_bold)
                    else:
                        col3 = Paragraph("", style_table_cell)
                        col4 = Paragraph("", style_table_cell)
                        
                    weights_data.append([col1, col2, col3, col4])
                    
                t_weights = Table(weights_data, colWidths=[180, 72, 180, 72])
                t_weights.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#0f766e")),
                    ("BACKGROUND", (2, 0), (3, 0), colors.HexColor("#0f766e")),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                    ("TOPPADDING", (0, 0), (-1, -1), 4.5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4.5),
                ]))
                story.append(t_weights)
                story.append(Spacer(1, 10))
            else:
                story.append(Paragraph("Allocation Strategy", style_h2))
                story.append(Paragraph("Equal Weight: Each strategy allocates 100% of the account equity (limitless concurrent mode).", style_body))
                story.append(Spacer(1, 10))
            
            explanation_text = (
                f"The <b>Conditional Drawdown at Risk (CVaR)</b> of <b>{cvar_pct:.2f}%</b> represents the average maximum "
                f"drawdown observed in the worst 5% (based on the <b>{conf_level:.1f}%</b> confidence level) of <b>{mc_paths}</b> "
                f"simulated Monte Carlo paths. Unlike simple historical drawdown which only analyzes the single historical sequence of trades, "
                f"the Bootstrap Resampling methodology randomly shuffles daily return sequences to simulate alternate 'parallel universe' paths. "
                f"This removes historical path-dependency and sequence risk, providing a highly conservative and mathematically robust estimation of "
                f"worst-case drawdown risk under stressful market conditions."
            )
            story.append(Paragraph(explanation_text, style_body))
            story.append(Spacer(1, 15))

        # --- Section 11: Actionable Suggestions ---
        story.append(Paragraph("Actionable Recommendations & Suggestions", style_h1))
        suggestions_list = self.generate_suggestions()
        for sug in suggestions_list:
            bullet_html = f"&bull; {sug}"
            story.append(Paragraph(bullet_html, style_suggestion))
            story.append(Spacer(1, 2))

        # Build Document
        doc.build(story, canvasmaker=NumberedCanvas)
        logger.info("PDF report compiled successfully.")

    # ==========================================================================
    # Excel Report Compiler
    # ==========================================================================

    def generate_excel(self, output_path: str) -> None:
        """
        Compiles a highly formatted multi-sheet Excel spreadsheet.

        Args:
            output_path: Destination file path for the Excel workbook.
        """
        logger.info(f"Generating Excel report: {output_path}")
        wb = openpyxl.Workbook()

        # Styles
        font_title = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
        font_section = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
        font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        font_bold = Font(name="Segoe UI", size=9, bold=True, color="1E293B")
        font_regular = Font(name="Segoe UI", size=9, color="334155")
        
        fill_title = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Navy
        fill_header_teal = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid") # Teal
        fill_header_grey = PatternFill(start_color="475569", end_color="475569", fill_type="solid") # Slate Grey
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Zebra Slate-50
        
        border_thin = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1")
        )

        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")

        # ----------------------------------------------------------------------
        # Sheet 1: Summary Dashboard
        # ----------------------------------------------------------------------
        ws = wb.active
        ws.title = "Summary Dashboard"
        ws.views.sheetView[0].showGridLines = True

        # Document Header Strip
        ws.merge_cells("A1:D2")
        title_cell = ws["A1"]
        title_cell.value = f"PORTFOLIO BACKTEST SUMMARY: {self.portfolio_name.upper()}"
        title_cell.font = font_title
        title_cell.fill = fill_title
        title_cell.alignment = align_center

        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 20

        # Performance Ratios Table
        ws["A4"] = "Portfolio Performance & Risk Ratios"
        ws["A4"].font = font_section
        
        perf_headers = ["Metric", "Value", "Ratio / Index", "Value"]
        for idx, h in enumerate(perf_headers, start=1):
            cell = ws.cell(row=5, column=idx, value=h)
            cell.font = font_header
            cell.fill = fill_header_teal
            cell.alignment = align_left

        initial_equity = self.results.get("initial_equity", 1000.0)
        ending_equity = self.results.get("ending_equity", 1000.0)
        net_profit = ending_equity - initial_equity
        net_profit_pct = (net_profit / initial_equity) if initial_equity > 0 else 0.0
        
        cagr = self.results.get("cagr", 0.0)
        max_dd = self.results.get("max_drawdown", 0.0) / 100.0

        perf_rows = [
            ("Initial Capital", initial_equity, "Sharpe Ratio", self.results.get("sharpe", 0.0)),
            ("Ending Capital", ending_equity, "Sortino Ratio", self.results.get("sortino", 0.0)),
            ("Net Profit ($)", net_profit, "Calmar Ratio", self.results.get("calmar", 0.0)),
            ("Net Profit (%)", net_profit_pct, "Ulcer Index", self.results.get("ulcer_index", 0.0)),
            ("CAGR", cagr, "Recovery Factor", self.results.get("recovery_factor", 0.0)),
            ("Max Drawdown (%)", max_dd, "Max Drawdown ($)", self.results.get("max_drawdown_cash", 0.0))
        ]

        number_formats = {
            "Initial Capital": "$#,##0.00",
            "Ending Capital": "$#,##0.00",
            "Net Profit ($)": "$#,##0.00;($#,##0.00);\"-\"",
            "Net Profit (%)": "0.00%",
            "CAGR": "0.00%",
            "Max Drawdown (%)": "0.00%",
            "Max Drawdown ($)": "$#,##0.00",
            "Sharpe Ratio": "0.00",
            "Sortino Ratio": "0.00",
            "Calmar Ratio": "0.00",
            "Ulcer Index": "0.00",
            "Recovery Factor": "0.00"
        }

        for row_idx, data in enumerate(perf_rows, start=6):
            ws.cell(row=row_idx, column=1, value=data[0])
            ws.cell(row=row_idx, column=2, value=data[1])
            ws.cell(row=row_idx, column=3, value=data[2])
            ws.cell(row=row_idx, column=4, value=data[3])

        # Apply formatting
        for r in range(6, 12):
            for c in [1, 3]:
                ws.cell(row=r, column=c).font = font_regular
                ws.cell(row=r, column=c).border = border_thin
            for c in [2, 4]:
                cell = ws.cell(row=r, column=c)
                cell.font = font_bold
                cell.border = border_thin
                lbl = ws.cell(row=r, column=c-1).value
                if lbl in number_formats:
                    cell.number_format = number_formats[lbl]

        # Trade Statistics Table
        ws["A14"] = "Portfolio Trade Statistics"
        ws["A14"].font = font_section

        t_stats = self.results.get("trade_statistics", {})
        trade_rows = [
            ("Total Trades Evaluated", t_stats.get("total_trades", len(self.trades)), "Average Trade PnL", t_stats.get("avg_trade", 0.0)),
            ("Executed Trades Count", t_stats.get("executed_trades", len(self.trades)), "Average Winning Trade", t_stats.get("avg_win", 0.0)),
            ("Skipped Trades", t_stats.get("skipped_trades", 0), "Average Losing Trade", t_stats.get("avg_loss", 0.0)),
            ("Win Rate (%)", t_stats.get("win_rate", 0.0) / 100.0, "Largest Winning Trade", t_stats.get("max_win", 0.0)),
            ("Profit Factor", t_stats.get("profit_factor", 0.0), "Largest Losing Trade", t_stats.get("max_loss", 0.0)),
            ("Expectancy", t_stats.get("expectancy", 0.0), "Max Streak (W/L)", f"{t_stats.get('max_consecutive_wins', 0)}W / {t_stats.get('max_consecutive_losses', 0)}L")
        ]

        trade_formats = {
            "Total Trades Evaluated": "#,##0",
            "Executed Trades Count": "#,##0",
            "Skipped Trades": "#,##0",
            "Win Rate (%)": "0.0%",
            "Profit Factor": "0.00",
            "Expectancy": "$#,##0.00",
            "Average Trade PnL": "$#,##0.00",
            "Average Winning Trade": "$#,##0.00",
            "Average Losing Trade": "$#,##0.00",
            "Largest Winning Trade": "$#,##0.00",
            "Largest Losing Trade": "$#,##0.00"
        }

        for idx, h in enumerate(perf_headers, start=1):
            cell = ws.cell(row=15, column=idx, value=h)
            cell.font = font_header
            cell.fill = fill_header_teal
            cell.alignment = align_left

        for row_idx, data in enumerate(trade_rows, start=16):
            ws.cell(row=row_idx, column=1, value=data[0])
            ws.cell(row=row_idx, column=2, value=data[1])
            ws.cell(row=row_idx, column=3, value=data[2])
            ws.cell(row=row_idx, column=4, value=data[3])

        for r in range(16, 22):
            for c in [1, 3]:
                ws.cell(row=r, column=c).font = font_regular
                ws.cell(row=r, column=c).border = border_thin
            for c in [2, 4]:
                cell = ws.cell(row=r, column=c)
                cell.font = font_bold
                cell.border = border_thin
                lbl = ws.cell(row=r, column=c-1).value
                if lbl in trade_formats:
                    cell.number_format = trade_formats[lbl]

        # ----------------------------------------------------------------------
        # Sheet 2: Monthly Matrix
        # ----------------------------------------------------------------------
        ws_monthly = wb.create_sheet(title="Monthly Matrix")
        ws_monthly.views.sheetView[0].showGridLines = True
        
        ws_monthly["A1"] = "Monthly Returns Matrix (%)"
        ws_monthly["A1"].font = font_section

        monthly_returns = self.results.get("monthly_returns", {})
        if monthly_returns:
            rows = []
            for key, val in monthly_returns.items():
                if isinstance(key, str):
                    parts = key.split("-")
                    year, month = int(parts[0]), int(parts[1])
                else:
                    year, month = key[0], key[1]
                rows.append({"Year": year, "Month": month, "Return": val})
            df = pd.DataFrame(rows)
            pivot_df = df.pivot(index="Year", columns="Month", values="Return")
            
            for m in range(1, 13):
                if m not in pivot_df.columns:
                    pivot_df[m] = np.nan
            pivot_df = pivot_df.reindex(columns=range(1, 13))
            
            headers = ["Year", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "YTD"]
            for c_idx, h in enumerate(headers, start=1):
                cell = ws_monthly.cell(row=3, column=c_idx, value=h)
                cell.font = font_header
                cell.fill = fill_header_teal
                cell.alignment = align_center

            for r_offset, yr in enumerate(pivot_df.index, start=4):
                ws_monthly.cell(row=r_offset, column=1, value=yr).font = font_bold
                ws_monthly.cell(row=r_offset, column=1).border = border_thin
                ytd_sum = 0.0
                for m in range(1, 13):
                    val = pivot_df.loc[yr, m]
                    cell = ws_monthly.cell(row=r_offset, column=m+1)
                    cell.border = border_thin
                    if not np.isnan(val):
                        # Convert to fraction for Excel % format
                        cell.value = val / 100.0
                        cell.number_format = "0.0%"
                        cell.font = font_bold if val >= 0 else font_regular
                        ytd_sum += val
                    else:
                        cell.value = "-"
                        cell.alignment = align_center
                        cell.font = font_regular
                        
                # YTD cell
                ytd_cell = ws_monthly.cell(row=r_offset, column=14, value=ytd_sum / 100.0)
                ytd_cell.font = font_bold
                ytd_cell.number_format = "0.0%"
                ytd_cell.border = border_thin

        # ----------------------------------------------------------------------
        # Sheet 3: Executed Trades Log
        # ----------------------------------------------------------------------
        ws_trades = wb.create_sheet(title="Executed Trades")
        ws_trades.views.sheetView[0].showGridLines = True
        
        ws_trades["A1"] = "Executed Trades Log"
        ws_trades["A1"].font = font_section

        trade_headers = [
            "Strategy Name", "Trade ID", "Side", "Entry Time", "Exit Time",
            "Entry Price", "Exit Price", "Contracts", "Position Value",
            "Commission", "Profit / Loss ($)", "Profit / Loss (%)"
        ]

        for col_idx, h in enumerate(trade_headers, start=1):
            cell = ws_trades.cell(row=3, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header_grey
            cell.alignment = align_left

        # Sort trades chronologically
        sorted_trades = sorted(self.trades, key=lambda t: t.exit_time or t.entry_time)
        for row_idx, t in enumerate(sorted_trades, start=4):
            ws_trades.cell(row=row_idx, column=1, value=t.strategy_name).alignment = align_left
            ws_trades.cell(row=row_idx, column=2, value=t.trade_id).alignment = align_center
            ws_trades.cell(row=row_idx, column=3, value=t.side).alignment = align_center
            
            ws_trades.cell(row=row_idx, column=4, value=t.entry_time.strftime("%Y-%m-%d %H:%M:%S") if t.entry_time else "-").alignment = align_center
            ws_trades.cell(row=row_idx, column=5, value=t.exit_time.strftime("%Y-%m-%d %H:%M:%S") if t.exit_time else "-").alignment = align_center
            
            ws_trades.cell(row=row_idx, column=6, value=t.entry_price).number_format = "#,##0.00"
            ws_trades.cell(row=row_idx, column=7, value=t.exit_price).number_format = "#,##0.00"
            ws_trades.cell(row=row_idx, column=8, value=t.contracts).number_format = "#,##0.0"
            ws_trades.cell(row=row_idx, column=9, value=t.position_value).number_format = "$#,##0.00"
            ws_trades.cell(row=row_idx, column=10, value=t.commission).number_format = "$#,##0.00"
            ws_trades.cell(row=row_idx, column=11, value=t.profit).number_format = "$#,##0.00;($#,##0.00);\"-\""
            ws_trades.cell(row=row_idx, column=12, value=t.profit_percent / 100.0).number_format = "0.00%"

            # Formatting fonts and zebra-striping
            zebra_fill = fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)
            for c in range(1, 13):
                cell = ws_trades.cell(row=row_idx, column=c)
                cell.font = font_regular
                cell.border = border_thin
                if zebra_fill.fill_type:
                    cell.fill = zebra_fill
                if c in [6, 7, 8, 9, 10, 11, 12]:
                    cell.alignment = align_right

        # ----------------------------------------------------------------------
        # Sheet 4: Margin Conflicts Log
        # ----------------------------------------------------------------------
        ws_conflicts = wb.create_sheet(title="Margin Conflicts")
        ws_conflicts.views.sheetView[0].showGridLines = True

        ws_conflicts["A1"] = "Margin Allocation Conflict Logs"
        ws_conflicts["A1"].font = font_section

        conflict_headers = [
            "Conflict Time", "Strategy Name", "Trade ID",
            "Required Margin", "Available Margin", "Status", "Skipped Trade PnL ($)"
        ]

        for col_idx, h in enumerate(conflict_headers, start=1):
            cell = ws_conflicts.cell(row=3, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header_grey
            cell.alignment = align_left

        for row_idx, log in enumerate(self.conflict_logs, start=4):
            dt_str = log["conflict_time"]
            if isinstance(dt_str, datetime):
                dt_str = dt_str.strftime("%Y-%m-%d %H:%M:%S")

            status_text = "Winner" if log.get("winner") else "Loser (Skipped)"
            
            skipped_pnl = ""
            if log.get("loser") and log.get("skipped_trade") is not None:
                skipped_pnl = getattr(log["skipped_trade"], "profit", 0.0)

            ws_conflicts.cell(row=row_idx, column=1, value=dt_str).alignment = align_center
            ws_conflicts.cell(row=row_idx, column=2, value=log.get("strategy", "Unknown")).alignment = align_left
            ws_conflicts.cell(row=row_idx, column=3, value=log.get("trade_id", "-")).alignment = align_center
            ws_conflicts.cell(row=row_idx, column=4, value=log.get("required_margin", 0.0)).number_format = "$#,##0.00"
            ws_conflicts.cell(row=row_idx, column=5, value=log.get("available_margin", 0.0)).number_format = "$#,##0.00"
            ws_conflicts.cell(row=row_idx, column=6, value=status_text).alignment = align_center
            
            pnl_cell = ws_conflicts.cell(row=row_idx, column=7, value=skipped_pnl)
            if skipped_pnl != "":
                pnl_cell.number_format = "$#,##0.00;($#,##0.00);\"-\""

            zebra_fill = fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)
            for c in range(1, 8):
                cell = ws_conflicts.cell(row=row_idx, column=c)
                cell.font = font_regular
                cell.border = border_thin
                if zebra_fill.fill_type:
                    cell.fill = zebra_fill
                if c in [4, 5, 7]:
                    cell.alignment = align_right

        # Auto-fit Column Widths across all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                for cell in col:
                    # Skip merged cells or title cells to prevent extreme column widths
                    if cell.row in [1, 2] and sheet.title in ["Summary Dashboard", "Executed Trades", "Margin Conflicts"]:
                        continue
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)

        wb.save(output_path)
        logger.info("Excel report compiled successfully.")

    # ==========================================================================
    # CSV Report Compiler
    # ==========================================================================

    def generate_csv(self, output_prefix: str) -> Dict[str, str]:
        """
        Compiles backtest results into standard CSV files.

        Args:
            output_prefix: Prefix path/name for the CSV files (e.g. "path/to/report").

        Returns:
            A dictionary mapping report names to their created CSV absolute file paths.
        """
        created_files = {}

        # 1. Summary Metrics CSV
        summary_path = f"{output_prefix}_summary.csv"
        logger.info(f"Generating Summary CSV: {summary_path}")
        
        flat_results = {}
        for k, v in self.results.items():
            if isinstance(v, (int, float, str, bool)) or v is None:
                flat_results[k] = v
        # Add nested trade stats
        trade_stats = self.results.get("trade_statistics", {})
        for k, v in trade_stats.items():
            if isinstance(v, (int, float, str, bool)):
                flat_results[f"trade_{k}"] = v
        
        df_summary = pd.DataFrame(list(flat_results.items()), columns=["MetricName", "Value"])
        df_summary.to_csv(summary_path, index=False)
        created_files["summary"] = os.path.abspath(summary_path)

        # 2. Monthly Returns CSV
        monthly_path = f"{output_prefix}_monthly_returns.csv"
        logger.info(f"Generating Monthly Returns CSV: {monthly_path}")
        monthly_returns = self.results.get("monthly_returns", {})
        df_monthly = pd.DataFrame(list(monthly_returns.items()), columns=["YearMonth", "ReturnPercent"])
        df_monthly.to_csv(monthly_path, index=False)
        created_files["monthly_returns"] = os.path.abspath(monthly_path)

        # 3. Executed Trades CSV
        trades_path = f"{output_prefix}_trades.csv"
        logger.info(f"Generating Trades CSV: {trades_path}")
        trade_rows = []
        for t in self.trades:
            trade_rows.append({
                "strategy_name": t.strategy_name,
                "trade_id": t.trade_id,
                "side": t.side,
                "entry_time": t.entry_time,
                "exit_time": t.exit_time,
                "entry_price": t.entry_price,
                "exit_price": t.exit_price,
                "contracts": t.contracts,
                "position_value": t.position_value,
                "commission": t.commission,
                "profit": t.profit,
                "profit_percent": t.profit_percent,
                "holding_time_seconds": t.holding_time.total_seconds() if t.holding_time else 0
            })
        df_trades = pd.DataFrame(trade_rows)
        df_trades.to_csv(trades_path, index=False)
        created_files["trades"] = os.path.abspath(trades_path)

        # 4. Conflicts CSV
        conflicts_path = f"{output_prefix}_conflicts.csv"
        logger.info(f"Generating Conflicts CSV: {conflicts_path}")
        conflict_rows = []
        for log in self.conflict_logs:
            skipped_pnl = ""
            if log.get("loser") and log.get("skipped_trade") is not None:
                skipped_pnl = getattr(log["skipped_trade"], "profit", 0.0)

            conflict_rows.append({
                "conflict_time": log.get("conflict_time"),
                "strategy_name": log.get("strategy"),
                "trade_id": log.get("trade_id"),
                "required_margin": log.get("required_margin"),
                "available_margin": log.get("available_margin"),
                "is_winner": log.get("winner"),
                "is_loser": log.get("loser"),
                "skipped_trade_profit": skipped_pnl
            })
        df_conflicts = pd.DataFrame(conflict_rows)
        df_conflicts.to_csv(conflicts_path, index=False)
        created_files["conflicts"] = os.path.abspath(conflicts_path)

        logger.info("CSV reports generated successfully.")
        return created_files
